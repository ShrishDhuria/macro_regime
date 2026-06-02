"""Build a parallel Excel workbook with named ranges, for interview walkthrough."""
from __future__ import annotations
import logging
from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

log = logging.getLogger(__name__)
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
SECTION_FONT = Font(bold=True, size=12)
PCT = "0.00%"


def _add_name(wb, name, attr_text):
    dn = DefinedName(name=name, attr_text=attr_text)
    try:
        wb.defined_names[name] = dn
    except (TypeError, AttributeError):
        wb.defined_names.append(dn)


def _autosize(ws):
    for col in ws.columns:
        try:
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)
        except Exception:
            pass


def _header_row(ws, row, values):
    for j, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=j, value=v)
        c.font = HEADER_FONT; c.fill = HEADER_FILL


def build_excel_workbook(portfolio_weights, portfolio_returns, var_table, es_table,
                         drawdown_info, regime_conditional_table, output_path: Path):
    wb = Workbook()
    ws = wb.active; ws.title = "Portfolio"
    ws["A1"] = "Multi-Asset Macro Portfolio"; ws["A1"].font = Font(bold=True, size=14)
    _header_row(ws, 3, ["Asset", "Weight"])
    row = 4
    for asset, w in portfolio_weights.items():
        ws.cell(row=row, column=1, value=asset)
        ws.cell(row=row, column=2, value=float(w)).number_format = "0%"
        row += 1
    ws.cell(row=row, column=1, value="Total").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"=SUM(B4:B{row-1})").number_format = "0%"
    _add_name(wb, "portfolio_weights", f"Portfolio!$B$4:$B${row-1}")

    ws.cell(row=row+3, column=1, value="Portfolio return statistics (weekly)").font = SECTION_FONT
    stats_rows = [
        ("Mean", portfolio_returns.mean(), PCT),
        ("Std dev", portfolio_returns.std(), PCT),
        ("Annualized vol", portfolio_returns.std() * np.sqrt(52), PCT),
        ("Skew", portfolio_returns.skew(), "0.000"),
        ("Excess kurtosis", portfolio_returns.kurtosis(), "0.000"),
        ("Min", portfolio_returns.min(), PCT),
        ("Max", portfolio_returns.max(), PCT),
        ("N observations", float(len(portfolio_returns)), "0"),
    ]
    for i, (name, val, fmt) in enumerate(stats_rows):
        r = row + 5 + i
        ws.cell(row=r, column=1, value=name).font = Font(bold=True)
        ws.cell(row=r, column=2, value=float(val)).number_format = fmt

    ws_ret = wb.create_sheet("Returns")
    _header_row(ws_ret, 1, ["Date", "Portfolio Return"])
    for i, (date, ret) in enumerate(portfolio_returns.items(), start=2):
        ws_ret.cell(row=i, column=1, value=date.date())
        ws_ret.cell(row=i, column=2, value=float(ret)).number_format = PCT
    _add_name(wb, "portfolio_returns", f"Returns!$B$2:$B${len(portfolio_returns)+1}")

    ws_r = wb.create_sheet("Risk Metrics")
    ws_r["A1"] = "Unconditional Risk Metrics"; ws_r["A1"].font = Font(bold=True, size=14)
    ws_r["A3"] = "Value-at-Risk (loss magnitude)"; ws_r["A3"].font = SECTION_FONT
    _header_row(ws_r, 4, ["Confidence", "Historical VaR", "Parametric VaR", "Cornish-Fisher VaR"])
    for i, rd in var_table.iterrows():
        r = 5 + i
        ws_r.cell(row=r, column=1, value=rd["confidence"])
        ws_r.cell(row=r, column=2, value=float(rd["historical_var"])).number_format = PCT
        ws_r.cell(row=r, column=3, value=float(rd["parametric_var"])).number_format = PCT
        ws_r.cell(row=r, column=4, value=float(rd["cornish_fisher_var"])).number_format = PCT
    es_start = 5 + len(var_table) + 2
    ws_r.cell(row=es_start, column=1, value="Expected Shortfall").font = SECTION_FONT
    _header_row(ws_r, es_start+1, ["Confidence", "Historical ES", "Parametric ES"])
    for i, rd in es_table.iterrows():
        r = es_start + 2 + i
        ws_r.cell(row=r, column=1, value=rd["confidence"])
        ws_r.cell(row=r, column=2, value=float(rd["historical_es"])).number_format = PCT
        ws_r.cell(row=r, column=3, value=float(rd["parametric_es"])).number_format = PCT
    dd_start = es_start + 2 + len(es_table) + 2
    ws_r.cell(row=dd_start, column=1, value="Drawdown").font = SECTION_FONT
    dd_rows = [
        ("Max drawdown", drawdown_info["max_drawdown"], PCT),
        ("Peak date", str(drawdown_info["peak_date"].date()), "@"),
        ("Trough date", str(drawdown_info["trough_date"].date()), "@"),
        ("Recovery date", str(drawdown_info["recovery_date"].date()) if drawdown_info["recovery_date"] else "Not recovered", "@"),
        ("Drawdown weeks", float(drawdown_info["drawdown_weeks"]), "0"),
        ("Recovery weeks", float(drawdown_info["recovery_weeks"]) if drawdown_info["recovery_weeks"] else None, "0"),
    ]
    for i, (name, val, fmt) in enumerate(dd_rows):
        r = dd_start + 1 + i
        ws_r.cell(row=r, column=1, value=name).font = Font(bold=True)
        cell = ws_r.cell(row=r, column=2, value=val if val is not None else "N/A")
        if isinstance(val, (int, float)):
            cell.number_format = fmt

    ws_g = wb.create_sheet("Regime Conditional")
    ws_g["A1"] = "Regime-Conditional Risk Metrics (3-state HMM)"; ws_g["A1"].font = Font(bold=True, size=14)
    ws_g["A2"] = "Same metrics computed separately on returns within each regime."; ws_g["A2"].font = Font(italic=True)
    _header_row(ws_g, 4, ["Metric"] + list(regime_conditional_table.columns))
    for i, (metric_name, rd) in enumerate(regime_conditional_table.iterrows()):
        r = 5 + i
        ws_g.cell(row=r, column=1, value=metric_name).font = Font(bold=True)
        for j, val in enumerate(rd.values):
            cell = ws_g.cell(row=r, column=2+j)
            if pd.notna(val):
                cell.value = float(val)
                if metric_name in ("Skew", "Excess kurtosis"):
                    cell.number_format = "0.000"
                elif metric_name == "N weeks":
                    cell.number_format = "0"
                else:
                    cell.number_format = PCT
            else:
                cell.value = "N/A"

    for ws_x in [ws, ws_ret, ws_r, ws_g]:
        _autosize(ws_x)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log.info(f"Saved Excel workbook to {output_path}")
    return output_path
